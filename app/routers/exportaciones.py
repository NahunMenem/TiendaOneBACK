from fastapi import APIRouter, Depends, Response
from sqlalchemy.orm import Session
from sqlalchemy import text
from app.deps import get_db
from app.routers.utils import get_current_user
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO

router = APIRouter(prefix="/exportar", tags=["Exportaciones"])

@router.get("/stock")
def exportar_stock(
    db: Session = Depends(get_db),
    user=Depends(get_current_user)
):
    productos = db.execute(text("""
        SELECT id, nombre, codigo_barras, num, color, bateria, condicion,
               stock, precio, precio_costo, precio_revendedor
        FROM productos_sj
        ORDER BY nombre
    """)).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Stock"

    headers = [
        "ID", "Nombre", "Código", "Núm", "Color", "Batería",
        "Condición", "Stock", "Precio Venta", "Precio Costo", "Precio Rev."
    ]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4B5563", end_color="4B5563", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    for fila in productos:
        ws.append(fila)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return Response(
        content=output.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=stock.xlsx"}
    )
