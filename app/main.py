# main.py
from fastapi import FastAPI, Depends, HTTPException, Request
from fastapi.responses import JSONResponse
from fastapi.middleware.cors import CORSMiddleware
from starlette.middleware.sessions import SessionMiddleware
from fastapi import UploadFile, File, HTTPException

import psycopg2
from psycopg2.extras import DictCursor
from datetime import datetime, timedelta
import pytz
import os

import cloudinary
import cloudinary.uploader

# =====================================================
# CONFIG
# =====================================================

import cloudinary
import cloudinary.uploader
import os

cloudinary.config(secure=True)

print("☁️ CLOUDINARY_URL:", os.getenv("CLOUDINARY_URL"))




from starlette.middleware.sessions import SessionMiddleware
from fastapi.middleware.cors import CORSMiddleware
from fastapi import FastAPI
import os

from fastapi import FastAPI
from fastapi.middleware.cors import CORSMiddleware

app = FastAPI(title="Backend SJ")

from fastapi.middleware.cors import CORSMiddleware

app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:3000",
        "http://127.0.0.1:3000",
        "https://tiendaone.vercel.app",
        "https://tiendauno.vercel.app",# si después deployás el front https://tiendauno.vercel.app/
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)





app.add_middleware(
    SessionMiddleware,
    secret_key=os.getenv("SECRET_KEY", "change-me"),
    same_site="none",     # ✅ necesario para cross-site
    https_only=True,      # ✅ Railway es HTTPS
)



# =====================================================
# DB
# =====================================================

def get_db():
    dsn = os.getenv("DATABASE_URL")
    if not dsn:
        raise RuntimeError("Falta DATABASE_URL")
    conn = psycopg2.connect(dsn, cursor_factory=DictCursor, sslmode="require")
    try:
        yield conn
    finally:
        conn.close()

# =====================================================
# INIT TABLES
# =====================================================

def crear_tabla_usuarios():
    conn = psycopg2.connect(os.getenv("DATABASE_URL"), cursor_factory=DictCursor, sslmode="require")
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS usuarios (
            id SERIAL PRIMARY KEY,
            username TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL,
            role TEXT NOT NULL DEFAULT 'user'
        )
    """)
    conn.commit()
    conn.close()

def crear_tabla_equipos():
    conn = psycopg2.connect(os.getenv("DATABASE_URL"), cursor_factory=DictCursor, sslmode="require")
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS equipos_tiendaone (
            id SERIAL PRIMARY KEY,
            tipo_reparacion TEXT NOT NULL,
            marca TEXT NOT NULL,
            modelo TEXT NOT NULL,
            tecnico TEXT NOT NULL,
            monto REAL NOT NULL,
            nombre_cliente TEXT NOT NULL,
            telefono TEXT NOT NULL,
            nro_orden TEXT NOT NULL,
            fecha TEXT NOT NULL,
            hora TEXT NOT NULL
        )
    """)
    conn.commit()
    conn.close()

crear_tabla_usuarios()
crear_tabla_equipos()

# =====================================================
# AUTH HELPERS
# =====================================================

def require_login(request: Request):
    if "username" not in request.session:
        raise HTTPException(status_code=401, detail="No autenticado")
    return request.session["username"]

# =====================================================
# AUTH
# =====================================================

@app.post("/login")
def login(data: dict, request: Request, db=Depends(get_db)):
    username = data.get("username")
    password = data.get("password")

    cur = db.cursor()
    cur.execute("SELECT * FROM usuarios WHERE username=%s", (username,))
    user = cur.fetchone()

    if not user or user["password"] != password:
        raise HTTPException(status_code=401, detail="Credenciales inválidas")

    request.session["username"] = user["username"]
    request.session["role"] = user["role"]

    return {"ok": True, "username": user["username"], "role": user["role"]}

@app.post("/logout")
def logout(request: Request):
    request.session.clear()
    return {"ok": True}

@app.get("/me")
def me(username=Depends(require_login), request: Request = None):
    return {
        "username": request.session.get("username"),
        "role": request.session.get("role"),
    }

# =====================================================
# CARRITO
# =====================================================

from fastapi import Depends, Request, HTTPException
from psycopg2.extras import DictCursor

@app.post("/carrito/limpiar")
def limpiar_carrito(request: Request):
    request.session["carrito"] = []
    return {"ok": True}

@app.post("/carrito/agregar")
def agregar_carrito(data: dict, request: Request, db=Depends(get_db)):
    if "carrito" not in request.session:
        request.session["carrito"] = []

    producto_id = data["producto_id"]
    cantidad = int(data.get("cantidad", 1))
    tipo_precio = data.get("tipo_precio", "venta")

    cur = db.cursor(cursor_factory=DictCursor)
    cur.execute("""
        SELECT
            id,
            nombre,
            stock,
            precio,
            precio_revendedor,
            moneda
        FROM productos_tiendaone
        WHERE id = %s
    """, (producto_id,))
    producto = cur.fetchone()

    if not producto:
        raise HTTPException(status_code=404, detail="Producto no encontrado")

    if producto["stock"] < cantidad:
        raise HTTPException(status_code=400, detail="Stock insuficiente")

    precio_usado = (
        float(producto["precio_revendedor"])
        if tipo_precio == "revendedor" and producto["precio_revendedor"]
        else float(producto["precio"])
    )

    request.session["carrito"].append({
        "id": producto["id"],
        "nombre": producto["nombre"],
        "precio": precio_usado,
        "cantidad": cantidad,
        "tipo_precio": tipo_precio,
        "moneda": producto["moneda"],   # 👈 CLAVE
    })

    cur.close()
    return {"ok": True}




from fastapi import Request, HTTPException

@app.post("/carrito/agregar-manual")
def agregar_manual(data: dict, request: Request):
    if "carrito" not in request.session:
        request.session["carrito"] = []

    moneda = data.get("moneda", "ARS")
    if moneda not in ("ARS", "USD"):
        raise HTTPException(
            status_code=400,
            detail="Moneda inválida. Use ARS o USD"
        )

    request.session["carrito"].append({
        "id": None,
        "nombre": data["nombre"],
        "precio": float(data["precio"]),
        "cantidad": int(data["cantidad"]),
        "tipo_precio": "manual",
        "moneda": moneda,   # 👈 CLAVE
    })

    return {"ok": True}



@app.post("/carrito/vaciar")
def vaciar_carrito(request: Request):
    request.session.pop("carrito", None)
    return {"ok": True}


from fastapi import Request

@app.get("/carrito")
def ver_carrito(request: Request):
    carrito = request.session.get("carrito", [])

    total = 0            # 👈 se mantiene para compatibilidad
    total_ars = 0
    total_usd = 0

    for i in carrito:
        subtotal = i["precio"] * i["cantidad"]
        total += subtotal

        if i.get("moneda") == "USD":
            total_usd += subtotal
        else:
            total_ars += subtotal

    return {
        "items": carrito,
        "total": total,              # 👈 NO se rompe nada existente
        "totales": {                 # 👈 nuevo, correcto
            "ARS": total_ars,
            "USD": total_usd
        }
    }


from fastapi import Request, HTTPException, Depends
from datetime import datetime

from fastapi import Depends, HTTPException, Request
from datetime import datetime

@app.post("/ventas/registrar")
async def registrar_venta(
    request: Request,
    db=Depends(get_db)
):
    data = await request.json()

    carrito = request.session.get("carrito", [])
    if not carrito:
        raise HTTPException(status_code=400, detail="Carrito vacío")

    dni_cliente = data.get("dni_cliente")
    pagos = data.get("pagos", [])

    if not dni_cliente:
        raise HTTPException(status_code=400, detail="Falta DNI del cliente")

    if not pagos:
        raise HTTPException(status_code=400, detail="Debe cargar al menos un pago")

    cur = db.cursor()
    fecha_actual = datetime.now()

    try:
        # =====================================================
        # 🧾 CREAR UNA SOLA VENTA (CABECERA)
        # =====================================================
        cur.execute("""
            INSERT INTO ventas_tiendaone (
                fecha,
                tipo_pago,
                dni_cliente,
                total
            )
            VALUES (%s, %s, %s, %s)
            RETURNING id
        """, (
            fecha_actual,
            "mixto",
            dni_cliente,
            0
        ))

        venta_id = cur.fetchone()[0]

        total_venta = 0

        # =====================================================
        # 📦 GUARDAR ÍTEMS (PRODUCTOS + MANUALES)
        # =====================================================
        for item in carrito:
            producto_id = item.get("id")
            cantidad = int(item["cantidad"])
            tipo_precio = item.get("tipo_precio") or "venta"

            # ---------- VENTA MANUAL ----------
            if producto_id is None:
                nombre_manual = item["nombre"]
                precio = float(item["precio"])
                moneda = item.get("moneda", "ARS")

                total = precio * cantidad
                total_venta += total

                cur.execute("""
                    INSERT INTO ventas_items_tiendaone (
                        venta_id,
                        producto_id,
                        nombre_manual,
                        cantidad,
                        precio_unitario,
                        moneda,
                        tipo_precio,
                        total
                    )
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
                """, (
                    venta_id,
                    None,
                    nombre_manual,
                    cantidad,
                    precio,
                    moneda,
                    tipo_precio,
                    total
                ))

            # ---------- PRODUCTO NORMAL ----------
            else:
                precio = float(item["precio"])
                moneda = item["moneda"]

                total = precio * cantidad
                total_venta += total

                cur.execute("""
                    INSERT INTO ventas_items_tiendaone (
                        venta_id,
                        producto_id,
                        cantidad,
                        precio_unitario,
                        moneda,
                        tipo_precio,
                        total
                    )
                    VALUES (%s,%s,%s,%s,%s,%s,%s)
                """, (
                    venta_id,
                    producto_id,
                    cantidad,
                    precio,
                    moneda,
                    tipo_precio,
                    total
                ))

                # 🔥 DESCONTAR STOCK
                cur.execute("""
                    UPDATE productos_tiendaone
                    SET stock = stock - %s
                    WHERE id = %s
                """, (cantidad, producto_id))

        # =====================================================
        # 💳 GUARDAR PAGOS (UNA SOLA VEZ)
        # =====================================================
        for p in pagos:
            cur.execute("""
                INSERT INTO pagos_tiendaone (
                    venta_id,
                    metodo,
                    moneda,
                    monto
                )
                VALUES (%s,%s,%s,%s)
            """, (
                venta_id,
                p["metodo"],
                p["moneda"],
                float(p["monto"])
            ))

        # =====================================================
        # 💰 ACTUALIZAR TOTAL DE LA VENTA
        # =====================================================
        cur.execute("""
            UPDATE ventas_tiendaone
            SET total = %s
            WHERE id = %s
        """, (total_venta, venta_id))

        db.commit()
        request.session["carrito"] = []
        return {"ok": True, "venta_id": venta_id}

    except Exception as e:
        db.rollback()
        print("ERROR REGISTRAR VENTA:", e)
        raise HTTPException(status_code=500, detail=str(e))

    finally:
        cur.close()






# =====================================================
# PRECIOS ACTUALIZADOS
# =====================================================

from fastapi import Depends, Request

@app.get("/api/carrito/precios_actualizados")
def precios_actualizados(
    tipo_precio: str = "venta",
    request: Request = None,
    db=Depends(get_db)
):
    carrito = request.session.get("carrito", [])
    cur = db.cursor()

    nuevos = []

    for item in carrito:
        # =========================
        # PRODUCTO DE BD
        # =========================
        if item["id"]:
            cur.execute("""
                SELECT precio, precio_revendedor, moneda
                FROM productos_tiendaone
                WHERE id=%s
            """, (item["id"],))

            p = cur.fetchone()

            precio = float(
                p["precio"] if tipo_precio == "venta" else p["precio_revendedor"]
            )

            nuevos.append({
                "id": item["id"],
                "nombre": item["nombre"],
                "cantidad": item["cantidad"],
                "precio": precio,
                "moneda": p["moneda"],        # 👈 CLAVE
                "tipo_precio": tipo_precio,
            })

        # =========================
        # ITEM MANUAL
        # =========================
        else:
            nuevos.append({
                "id": None,
                "nombre": item["nombre"],
                "cantidad": item["cantidad"],
                "precio": float(item["precio"]),
                "moneda": item.get("moneda", "ARS"),  # 👈 DEFAULT SEGURO
                "tipo_precio": tipo_precio,
            })

    cur.close()
    return nuevos


# =====================================================
# PRODUCTOS MÁS VENDIDOS
# =====================================================

# =====================================================
# PRODUCTOS MÁS VENDIDOS (TOP 15)
# =====================================================

@app.get("/productos_mas_vendidos")
def productos_mas_vendidos(db=Depends(get_db)):
    cur = db.cursor()

    # Total general de unidades vendidas
    cur.execute("""
        SELECT COALESCE(SUM(cantidad), 0) AS total
        FROM ventas_tiendaone
        WHERE producto_id IS NOT NULL
    """)
    total = cur.fetchone()["total"]

    # Top 15 productos reales
    cur.execute("""
        SELECT
            p.nombre,
            p.precio,
            SUM(v.cantidad) AS unidades
        FROM ventas_tiendaone v
        JOIN productos_tiendaone p ON p.id = v.producto_id
        WHERE v.producto_id IS NOT NULL
        GROUP BY p.id, p.nombre, p.precio
        ORDER BY unidades DESC
        LIMIT 15
    """)
    productos = cur.fetchall()

    return {
        "total_ventas": int(total),
        "productos": [
            {
                "nombre": p["nombre"],
                "precio": float(p["precio"]),
                "unidades": int(p["unidades"]),
                "porcentaje": round((p["unidades"] / total) * 100, 2) if total else 0,
            }
            for p in productos
        ],
    }


# =====================================================
# PRODUCTOS MÁS VENDIDOS (DETALLADO)
# =====================================================

@app.get("/productos_mas_vendidos/detalle")
def productos_mas_vendidos_detalle(db=Depends(get_db)):
    cur = db.cursor()

    # Total general
    cur.execute("""
        SELECT COALESCE(SUM(cantidad), 0) AS total
        FROM ventas_tiendaone
        WHERE producto_id IS NOT NULL
    """)
    total_ventas = cur.fetchone()["total"]

    # Top 15
    cur.execute("""
        SELECT
            p.nombre,
            p.precio,
            SUM(v.cantidad) AS unidades
        FROM ventas_tiendaone v
        JOIN productos_tiendaone p ON p.id = v.producto_id
        WHERE v.producto_id IS NOT NULL
        GROUP BY p.id, p.nombre, p.precio
        ORDER BY unidades DESC
        LIMIT 15
    """)
    productos = cur.fetchall()

    return {
        "total_ventas": int(total_ventas),
        "productos": [
            {
                "nombre": p["nombre"],
                "precio": float(p["precio"]),
                "unidades": int(p["unidades"]),
                "porcentaje": round(
                    (p["unidades"] / total_ventas) * 100, 2
                ) if total_ventas else 0,
            }
            for p in productos
        ],
    }


# =====================================================
# PRODUCTOS POR AGOTARSE
# =====================================================

# =====================================================
# PRODUCTOS POR AGOTARSE (PAGINADO)
# =====================================================

@app.get("/productos_por_agotarse")
def productos_por_agotarse(
    page: int = 1,
    db=Depends(get_db)
):
    cur = db.cursor()

    limit = 20
    offset = (page - 1) * limit

    # Total para paginación
    cur.execute("""
        SELECT COUNT(*) AS total
        FROM productos_tiendaone
        WHERE stock <= 30
    """)
    total = cur.fetchone()["total"]

    # Datos paginados
    cur.execute("""
        SELECT id, nombre, codigo_barras, stock, precio, precio_costo
        FROM productos_tiendaone
        WHERE stock <= 30
        ORDER BY stock ASC
        LIMIT %s OFFSET %s
    """, (limit, offset))

    productos = cur.fetchall()

    return {
        "page": page,
        "limit": limit,
        "total": total,
        "productos": [
            {
                "id": p["id"],
                "nombre": p["nombre"],
                "codigo_barras": p["codigo_barras"],
                "stock": p["stock"],
                "precio": float(p["precio"]),
                "precio_costo": float(p["precio_costo"]) if p["precio_costo"] else 0,
            }
            for p in productos
        ],
    }


# =====================================================
# ÚLTIMAS VENTAS Y REPARACIONES
# =====================================================
@app.get("/ventas/buscar")
def buscar_productos(busqueda: str, db=Depends(get_db)):
    cur = db.cursor(cursor_factory=DictCursor)
    cur.execute("""
        SELECT id, nombre, codigo_barras, num, stock, precio, precio_revendedor
        FROM productos_tiendaone
        WHERE codigo_barras = %s
           OR nombre ILIKE %s
           OR num ILIKE %s
        ORDER BY nombre
    """, (busqueda, f"%{busqueda}%", f"%{busqueda}%"))

    return cur.fetchall()



from openpyxl import Workbook
from io import BytesIO
from fastapi.responses import StreamingResponse

@app.get("/ultimas_ventas")
def ultimas_ventas(
    fecha_desde: str | None = None,
    fecha_hasta: str | None = None,
    exportar: bool = False,
    db=Depends(get_db)
):
    cur = db.cursor()

    argentina_tz = pytz.timezone("America/Argentina/Buenos_Aires")
    hoy = datetime.now(argentina_tz).strftime("%Y-%m-%d")

    fecha_desde = fecha_desde or hoy
    fecha_hasta = fecha_hasta or hoy

    # =============================
    # VENTAS TIENDAONE (NORMAL + MANUAL)
    # =============================
    cur.execute("""
        SELECT 
            v.id AS venta_id,

            COALESCE(p.nombre, v.nombre_manual) AS nombre_producto,

            COALESCE(p.num, '-') AS num,

            v.cantidad,

            COALESCE(v.precio_unitario, v.precio_manual) AS precio_unitario,

            v.total,
            v.fecha,
            v.tipo_pago,
            v.dni_cliente,
            v.tipo_precio
        FROM ventas_tiendaone v
        LEFT JOIN productos_tiendaone p ON v.producto_id = p.id
        WHERE DATE(v.fecha) BETWEEN %s AND %s
        ORDER BY v.fecha DESC
    """, (fecha_desde, fecha_hasta))

    ventas = cur.fetchall()

    # =============================
    # TOTALES POR MÉTODO DE PAGO
    # =============================
    total_ventas_por_pago = {}
    for v in ventas:
        total_ventas_por_pago[v["tipo_pago"]] = (
            total_ventas_por_pago.get(v["tipo_pago"], 0) + (v["total"] or 0)
        )

    # =============================
    # EXPORTAR EXCEL
    # =============================
    if exportar:
        wb = Workbook()
        ws = wb.active
        ws.title = "Ventas"

        ws.append([
            "ID Venta", "Producto", "Cantidad", "Núm",
            "Precio Unitario", "Total", "Fecha",
            "Tipo de Pago", "DNI Cliente", "Tipo Precio"
        ])

        for v in ventas:
            ws.append([
                v["venta_id"],
                v["nombre_producto"],
                v["cantidad"],
                v["num"],
                float(v["precio_unitario"] or 0),
                float(v["total"] or 0),
                v["fecha"].strftime("%Y-%m-%d %H:%M:%S") if v["fecha"] else "",
                v["tipo_pago"],
                v["dni_cliente"] or "",
                v["tipo_precio"].capitalize() if v["tipo_precio"] else "",
            ])

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"ventas_{fecha_desde}_a_{fecha_hasta}.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    return {
        "ventas": ventas,
        "totales_ventas_por_pago": total_ventas_por_pago,
        "fecha_desde": fecha_desde,
        "fecha_hasta": fecha_hasta,
    }


# =====================================================
# ANULAR VENTA
# =====================================================

@app.delete("/anular_venta/{venta_id}")
def anular_venta(venta_id: int, db=Depends(get_db)):
    cur = db.cursor()

    cur.execute("SELECT id FROM ventas_tiendaone WHERE id=%s", (venta_id,))
    if not cur.fetchone():
        raise HTTPException(status_code=404, detail="Venta no encontrada")

    cur.execute("DELETE FROM ventas_tiendaone WHERE id=%s", (venta_id,))
    db.commit()

    return {"success": True}

# =====================================================
# ANULAR REPARACIÓN
# =====================================================

@app.delete("/anular_reparacion/{reparacion_id}")
def anular_reparacion(reparacion_id: int, db=Depends(get_db)):
    cur = db.cursor()

    cur.execute("SELECT id FROM reparaciones_tiendaone WHERE id=%s", (reparacion_id,))
    if not cur.fetchone():
        raise HTTPException(status_code=404, detail="Reparación no encontrada")

    cur.execute("DELETE FROM reparaciones_tiendaone WHERE id=%s", (reparacion_id,))
    db.commit()

    return {"success": True}

# =====================================================
# EGRESOS
# =====================================================

from datetime import datetime

@app.get("/egresos")
def listar_egresos(
    fecha_desde: str | None = None,
    fecha_hasta: str | None = None,
    db=Depends(get_db)
):
    cur = db.cursor()

    if fecha_desde and fecha_hasta:
        cur.execute("""
            SELECT id, fecha, monto, descripcion, tipo_pago
            FROM egresos_tiendaone
            WHERE fecha BETWEEN %s AND %s
            ORDER BY fecha DESC
        """, (
            f"{fecha_desde} 00:00:00",
            f"{fecha_hasta} 23:59:59",
        ))
    else:
        cur.execute("""
            SELECT id, fecha, monto, descripcion, tipo_pago
            FROM egresos_tiendaone
            ORDER BY fecha DESC
        """)

    egresos = cur.fetchall()

    return [
        {
            "id": e["id"],
            "fecha": e["fecha"],
            "monto": float(e["monto"]),
            "descripcion": e["descripcion"],
            "tipo_pago": e["tipo_pago"],
        }
        for e in egresos
    ]


@app.post("/egresos")
def agregar_egreso(data: dict, db=Depends(get_db)):
    cur = db.cursor()
    cur.execute("""
        INSERT INTO egresos_tiendaone (fecha, monto, descripcion, tipo_pago)
        VALUES (%s, %s, %s, %s)
    """, (
        data["fecha"],
        float(data["monto"]),
        data["descripcion"],
        data["tipo_pago"],
    ))
    db.commit()
    return {"ok": True}

@app.delete("/egresos/{egreso_id}")
def eliminar_egreso(egreso_id: int, db=Depends(get_db)):
    cur = db.cursor()
    cur.execute("DELETE FROM egresos_tiendaone WHERE id=%s", (egreso_id,))
    db.commit()
    return {"ok": True}

# =====================================================
# DASHBOARD
# =====================================================

@app.get("/dashboard")
def dashboard(
    fecha_desde: str | None = None,
    fecha_hasta: str | None = None,
    db=Depends(get_db)
):
    cur = db.cursor(cursor_factory=DictCursor)

    hoy = datetime.now().strftime("%Y-%m-%d")
    fecha_desde = fecha_desde or hoy
    fecha_hasta = fecha_hasta or hoy

    # =====================================================
    # 💳 INGRESOS REALES DESDE PAGOS (POR MONEDA)
    # =====================================================
    cur.execute("""
        SELECT
            pg.moneda,
            SUM(pg.monto) AS total
        FROM pagos_tiendaone pg
        JOIN ventas_tiendaone v ON v.id = pg.venta_id
        WHERE DATE(v.fecha) BETWEEN %s AND %s
        GROUP BY pg.moneda
    """, (fecha_desde, fecha_hasta))

    ingresos_por_moneda = {
        r["moneda"]: float(r["total"] or 0)
        for r in cur.fetchall()
    }

    total_ingresos = sum(ingresos_por_moneda.values())

    # =====================================================
    # 🔧 REPARACIONES (ARS)
    # =====================================================
    cur.execute("""
        SELECT SUM(precio) AS total
        FROM reparaciones_tiendaone
        WHERE DATE(fecha) BETWEEN %s AND %s
    """, (fecha_desde, fecha_hasta))

    total_reparaciones = float(cur.fetchone()["total"] or 0)

    # =====================================================
    # 📉 EGRESOS (ARS)
    # =====================================================
    cur.execute("""
        SELECT SUM(monto) AS total
        FROM egresos_tiendaone
        WHERE DATE(fecha) BETWEEN %s AND %s
    """, (fecha_desde, fecha_hasta))

    total_egresos = float(cur.fetchone()["total"] or 0)

    # =====================================================
    # 📦 COSTO DE PRODUCTOS VENDIDOS (POR MONEDA) ✅ CORRECTO
    # =====================================================
    cur.execute("""
        SELECT
            p.moneda,
            SUM(vi.cantidad * p.precio_costo) AS total
        FROM ventas_items_tiendaone vi
        JOIN ventas_tiendaone v ON v.id = vi.venta_id
        JOIN productos_tiendaone p ON p.id = vi.producto_id
        WHERE DATE(v.fecha) BETWEEN %s AND %s
        GROUP BY p.moneda
    """, (fecha_desde, fecha_hasta))

    costo_por_moneda = {
        r["moneda"]: float(r["total"] or 0)
        for r in cur.fetchall()
    }

    total_costo = sum(costo_por_moneda.values())

    # =====================================================
    # 🧮 GANANCIA POR MONEDA (CORRECTA)
    # =====================================================
    ganancia_por_moneda = {}

    for moneda, ingresos in ingresos_por_moneda.items():
        costo = costo_por_moneda.get(moneda, 0)
        egreso = total_egresos if moneda == "ARS" else 0
        reparaciones = total_reparaciones if moneda == "ARS" else 0

        ganancia_por_moneda[moneda] = ingresos - costo - egreso + reparaciones

    ganancia_total = sum(ganancia_por_moneda.values())

    # =====================================================
    # 📊 DISTRIBUCIÓN (LEGACY, NO SE ROMPE)
    # =====================================================
    cur.execute("""
        SELECT 'Productos' AS tipo, SUM(total) AS total
        FROM ventas_tiendaone
        WHERE DATE(fecha) BETWEEN %s AND %s

        UNION ALL

        SELECT 'Reparaciones' AS tipo, SUM(precio)
        FROM reparaciones_tiendaone
        WHERE DATE(fecha) BETWEEN %s AND %s
    """, (fecha_desde, fecha_hasta, fecha_desde, fecha_hasta))

    distribucion = cur.fetchall()

    cur.close()

    return {
        # ================== EXISTENTE ==================
        "fecha_desde": fecha_desde,
        "fecha_hasta": fecha_hasta,
        "total_ventas": total_ingresos,
        "total_ventas_reparaciones": total_reparaciones,
        "total_egresos": total_egresos,
        "total_costo": total_costo,
        "ganancia": ganancia_total,
        "distribucion_ventas": [
            {"tipo": d["tipo"], "total": float(d["total"] or 0)}
            for d in distribucion
        ],

        # ================== NUEVO / CORRECTO ==================
        "ingresos_por_moneda": ingresos_por_moneda,
        "costo_por_moneda": costo_por_moneda,
        "ganancia_por_moneda": ganancia_por_moneda
    }




# =====================================================
# RESUMEN SEMANAL
# =====================================================

@app.get("/resumen_semanal")
def resumen_semanal(db=Depends(get_db)):
    hoy = datetime.now()
    inicio_semana = hoy - timedelta(days=hoy.weekday())
    inicio_semana_str = inicio_semana.strftime("%Y-%m-%d")

    cur = db.cursor()
    cur.execute("""
        SELECT tipo_pago, SUM(total) AS total
        FROM ventas_tiendaone
        WHERE fecha >= %s
        GROUP BY tipo_pago
    """, (inicio_semana_str,))

    resumen = cur.fetchall()

    return [
        {
            "tipo_pago": r["tipo_pago"],
            "total": r["total"] or 0,
        }
        for r in resumen
    ]

# =====================================================
# CAJA
# =====================================================

@app.get("/caja")
def caja(
    fecha_desde: str | None = None,
    fecha_hasta: str | None = None,
    db=Depends(get_db)
):
    argentina_tz = pytz.timezone("America/Argentina/Buenos_Aires")
    hoy = datetime.now(argentina_tz).date()

    fecha_desde = fecha_desde or hoy.strftime("%Y-%m-%d")
    fecha_hasta = fecha_hasta or hoy.strftime("%Y-%m-%d")

    desde_dt = datetime.fromisoformat(fecha_desde)
    hasta_dt = datetime.fromisoformat(fecha_hasta) + timedelta(days=1)

    cur = db.cursor(cursor_factory=DictCursor)

    # =====================================================
    # 💳 INGRESOS REALES (VENTAS + MANUALES) DESDE PAGOS
    # =====================================================
    cur.execute("""
        SELECT
            pg.metodo AS tipo_pago,
            pg.moneda,
            SUM(pg.monto) AS total
        FROM pagos_tiendaone pg
        JOIN ventas_tiendaone v ON v.id = pg.venta_id
        WHERE v.fecha BETWEEN %s AND %s
        GROUP BY pg.metodo, pg.moneda
    """, (desde_dt, hasta_dt))

    pagos = cur.fetchall()

    # =====================================================
    # 🔧 REPARACIONES (ARS)
    # =====================================================
    cur.execute("""
        SELECT
            tipo_pago,
            SUM(cantidad * precio) AS total
        FROM reparaciones_tiendaone
        WHERE fecha BETWEEN %s AND %s
        GROUP BY tipo_pago
    """, (desde_dt, hasta_dt))

    reparaciones = cur.fetchall()

    # =====================================================
    # 📉 EGRESOS (ARS)
    # =====================================================
    cur.execute("""
        SELECT
            tipo_pago,
            SUM(monto) AS total
        FROM egresos_tiendaone
        WHERE fecha BETWEEN %s AND %s
        GROUP BY tipo_pago
    """, (desde_dt, hasta_dt))

    egresos = cur.fetchall()

    # =====================================================
    # 🧮 CONSOLIDACIÓN (COMPATIBLE + POR MONEDA)
    # =====================================================
    total_por_pago = {}
    total_por_pago_moneda = {}

    # INGRESOS (pagos reales)
    for p in pagos:
        tipo = p["tipo_pago"]
        moneda = p["moneda"]
        total = float(p["total"] or 0)

        # legacy
        total_por_pago[tipo] = total_por_pago.get(tipo, 0) + total

        # por moneda
        total_por_pago_moneda.setdefault(tipo, {})
        total_por_pago_moneda[tipo][moneda] = (
            total_por_pago_moneda[tipo].get(moneda, 0) + total
        )

    # REPARACIONES (ARS)
    for r in reparaciones:
        tipo = r["tipo_pago"]
        total = float(r["total"] or 0)

        total_por_pago[tipo] = total_por_pago.get(tipo, 0) + total

        total_por_pago_moneda.setdefault(tipo, {})
        total_por_pago_moneda[tipo]["ARS"] = (
            total_por_pago_moneda[tipo].get("ARS", 0) + total
        )

    # EGRESOS (ARS)
    egresos_por_pago = {
        e["tipo_pago"]: float(e["total"] or 0)
        for e in egresos
    }

    # NETO LEGACY
    neto_por_pago = {
        tipo: total - egresos_por_pago.get(tipo, 0)
        for tipo, total in total_por_pago.items()
    }

    # NETO POR MONEDA (CORRECTO)
    neto_por_pago_moneda = {}

    for tipo, monedas in total_por_pago_moneda.items():
        neto_por_pago_moneda[tipo] = {}
        for moneda, total in monedas.items():
            egreso = egresos_por_pago.get(tipo, 0) if moneda == "ARS" else 0
            neto_por_pago_moneda[tipo][moneda] = total - egreso

    cur.close()

    return {
        "fecha_desde": fecha_desde,
        "fecha_hasta": fecha_hasta,

        # 👈 compatibilidad vieja
        "neto_por_pago": neto_por_pago,

        # 👈 caja correcta separada por moneda
        "neto_por_pago_moneda": neto_por_pago_moneda
    }


# =====================================================
# REPARACIONES
# =====================================================

import unicodedata

def normalizar(texto: str):
    return unicodedata.normalize("NFKD", texto).encode("ascii", "ignore").decode().lower().strip()


# =====================================================
# ELIMINAR REPARACIÓN (EQUIPO)
# =====================================================

@app.delete("/reparaciones/{id}")
def eliminar_reparacion(id: int, db=Depends(get_db)):
    cur = db.cursor()
    cur.execute(
        "DELETE FROM reparaciones_tiendaone WHERE id = %s",
        (id,)
    )
    db.commit()
    cur.close()
    return {"ok": True}

# =====================================================
# ACTUALIZAR ESTADO DE REPARACIÓN
# =====================================================

@app.post("/reparaciones/actualizar_estado")
def actualizar_estado(data: dict, db=Depends(get_db)):
    nro_orden = data["nro_orden"]
    estado = data["estado"]

    cur = db.cursor()
    cur.execute("""
        UPDATE equipos_tiendaone
        SET estado = %s
        WHERE nro_orden = %s
    """, (estado, nro_orden))
    db.commit()

    return {"success": True}

# =====================================================
# MERCADERÍA FALLADA
# =====================================================

@app.get("/mercaderia_fallada")
def listar_mercaderia_fallada(db=Depends(get_db)):
    cur = db.cursor()
    cur.execute("""
        SELECT mf.id, p.nombre, mf.cantidad, mf.fecha, mf.descripcion
        FROM mercaderia_fallada mf
        JOIN productos_tiendaone p ON mf.producto_id = p.id
        ORDER BY mf.fecha DESC
    """)
    return cur.fetchall()


@app.post("/mercaderia_fallada")
def registrar_mercaderia_fallada(data: dict, db=Depends(get_db)):
    producto_id = data["producto_id"]
    cantidad = int(data["cantidad"])
    descripcion = data["descripcion"]
    fecha = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    cur = db.cursor()
    cur.execute("SELECT stock FROM productos_tiendaone WHERE id=%s", (producto_id,))
    prod = cur.fetchone()

    if not prod or prod["stock"] < cantidad:
        raise HTTPException(status_code=400, detail="Stock insuficiente")

    cur.execute("""
        INSERT INTO mercaderia_fallada_tiendaone (producto_id, cantidad, fecha, descripcion)
        VALUES (%s, %s, %s, %s)
    """, (producto_id, cantidad, fecha, descripcion))

    cur.execute("""
        UPDATE productos_tiendaone
        SET stock = stock - %s
        WHERE id = %s
    """, (cantidad, producto_id))

    db.commit()
    return {"success": True}

# =====================================================
# CATEGORÍAS
# =====================================================

def crear_tabla_categorias():
    conn = psycopg2.connect(os.getenv("DATABASE_URL"), cursor_factory=DictCursor, sslmode="require")
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS categorias_tiendaone (
            id SERIAL PRIMARY KEY,
            nombre TEXT UNIQUE NOT NULL
        )
    """)
    conn.commit()
    conn.close()

crear_tabla_categorias()


@app.get("/categorias")
def listar_categorias(db=Depends(get_db)):
    cur = db.cursor()
    cur.execute("SELECT nombre FROM categorias_tiendaone ORDER BY nombre")
    return [r["nombre"] for r in cur.fetchall()]


@app.post("/categorias")
def agregar_categoria(data: dict, db=Depends(get_db)):
    cur = db.cursor()
    cur.execute(
        "INSERT INTO categorias_tiendaone (nombre) VALUES (%s) ON CONFLICT DO NOTHING",
        (data["nombre"].strip(),)
    )
    db.commit()
    return {"ok": True}


@app.delete("/categorias/{nombre}")
def eliminar_categoria(nombre: str, db=Depends(get_db)):
    cur = db.cursor()
    cur.execute("DELETE FROM categorias_tiendaone WHERE nombre=%s", (nombre,))
    db.commit()
    return {"ok": True}

# =====================================================
# PRODUCTOS / STOCK
# =====================================================
@app.delete("/productos/{id}")
def eliminar_producto(id: int, db=Depends(get_db)):
    try:
        cur = db.cursor()
        cur.execute("DELETE FROM productos_tiendaone WHERE id = %s", (id,))
        db.commit()
        return {"ok": True}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


from psycopg2.extras import DictCursor
from fastapi import Depends

@app.get("/productos")
def listar_productos(
    busqueda: str | None = None,
    page: int = 1,
    limit: int = 20,
    db=Depends(get_db)
):
    offset = (page - 1) * limit
    cur = db.cursor(cursor_factory=DictCursor)

    if busqueda:
        cur.execute("""
            SELECT
                id,
                nombre,
                codigo_barras,
                stock,
                precio,
                moneda,
                precio_costo,
                precio_revendedor,
                categoria,
                foto_url,
                num,
                color,
                bateria,
                condicion
            FROM productos_tiendaone
            WHERE nombre ILIKE %s
               OR codigo_barras ILIKE %s
               OR num ILIKE %s
            ORDER BY id DESC
            LIMIT %s OFFSET %s
        """, (
            f"%{busqueda}%",
            f"%{busqueda}%",
            f"%{busqueda}%",
            limit,
            offset
        ))
    else:
        cur.execute("""
            SELECT
                id,
                nombre,
                codigo_barras,
                stock,
                precio,
                moneda,
                precio_costo,
                precio_revendedor,
                categoria,
                foto_url,
                num,
                color,
                bateria,
                condicion
            FROM productos_tiendaone
            ORDER BY id DESC
            LIMIT %s OFFSET %s
        """, (limit, offset))

    rows = cur.fetchall()

    # total para paginación
    if busqueda:
        cur.execute("""
            SELECT COUNT(*)
            FROM productos_tiendaone
            WHERE nombre ILIKE %s
               OR codigo_barras ILIKE %s
               OR num ILIKE %s
        """, (
            f"%{busqueda}%",
            f"%{busqueda}%",
            f"%{busqueda}%"
        ))
    else:
        cur.execute("SELECT COUNT(*) FROM productos_tiendaone")

    total = cur.fetchone()[0]

    cur.close()

    return {
        "items": [
            {
                "id": r["id"],
                "nombre": r["nombre"],
                "codigo_barras": r["codigo_barras"],
                "stock": r["stock"],
                "precio": float(r["precio"]) if r["precio"] is not None else 0,
                "moneda": r["moneda"],                 # 👈 CLAVE
                "precio_costo": float(r["precio_costo"]) if r["precio_costo"] else None,
                "precio_revendedor": float(r["precio_revendedor"]) if r["precio_revendedor"] else None,
                "categoria": r["categoria"],
                "foto_url": r["foto_url"],
                "num": r["num"],
                "color": r["color"],
                "bateria": r["bateria"],
                "condicion": r["condicion"],
            }
            for r in rows
        ],
        "total": total,
        "page": page,
        "limit": limit
    }



# ==========================
# SUBIDA DE IMÁGENES (Cloudinary)
# ==========================
from fastapi import UploadFile, File, HTTPException
import cloudinary
import cloudinary.uploader
import traceback

@app.post("/upload-imagen")
async def upload_imagen(file: UploadFile = File(...)):
    try:
        print("📷 Archivo recibido:", file.filename, file.content_type)

        result = cloudinary.uploader.upload(
            file.file,
            folder="productos",
            resource_type="image"
        )

        return {
            "ok": True,
            "url": result["secure_url"]
        }

    except Exception as e:
        print("❌ ERROR CLOUDINARY", e)
        raise HTTPException(status_code=500, detail=str(e))



from psycopg2.extras import DictCursor
from fastapi import Depends, HTTPException

@app.post("/productos")
def agregar_producto(data: dict, db=Depends(get_db)):
    try:
        cur = db.cursor(cursor_factory=DictCursor)

        # 🔒 Validación simple de moneda
        moneda = data.get("moneda", "ARS")
        if moneda not in ("ARS", "USD"):
            raise HTTPException(
                status_code=400,
                detail="Moneda inválida. Use ARS o USD"
            )

        cur.execute("""
            INSERT INTO productos_tiendaone (
                nombre,
                codigo_barras,
                stock,
                precio,
                moneda,
                precio_costo,
                foto_url,
                categoria,
                num,
                color,
                bateria,
                precio_revendedor,
                condicion
            ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, (
            data["nombre"].upper(),
            data["codigo_barras"],
            int(data["stock"]),
            float(data["precio"]),
            moneda,                               # 👈 CLAVE
            float(data["precio_costo"]),
            data.get("foto_url"),
            data.get("categoria"),
            data.get("num"),
            data.get("color"),
            data.get("bateria"),
            data.get("precio_revendedor"),
            data.get("condicion"),
        ))

        db.commit()
        cur.close()
        return {"ok": True}

    except KeyError as e:
        db.rollback()
        raise HTTPException(
            status_code=400,
            detail=f"Falta el campo obligatorio: {e}"
        )

    except HTTPException:
        db.rollback()
        raise

    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=500,
            detail=str(e)
        )



from fastapi import Depends, HTTPException

@app.put("/productos/{id}")
def editar_producto(id: int, data: dict, db=Depends(get_db)):
    try:
        cur = db.cursor()

        # 🔒 Validación simple de moneda
        moneda = data.get("moneda", "ARS")
        if moneda not in ("ARS", "USD"):
            raise HTTPException(
                status_code=400,
                detail="Moneda inválida. Use ARS o USD"
            )

        cur.execute("""
            UPDATE productos_tiendaone
            SET nombre=%s,
                codigo_barras=%s,
                stock=%s,
                precio=%s,
                moneda=%s,
                precio_costo=%s,
                foto_url=%s,
                categoria=%s,
                num=%s,
                color=%s,
                bateria=%s,
                precio_revendedor=%s,
                condicion=%s
            WHERE id=%s
        """, (
            data["nombre"].upper(),
            data["codigo_barras"],
            int(data["stock"]),
            float(data["precio"]),
            moneda,                         # 👈 CLAVE
            float(data["precio_costo"]),
            data.get("foto_url"),
            data.get("categoria"),
            data.get("num"),
            data.get("color"),
            data.get("bateria"),
            data.get("precio_revendedor"),
            data.get("condicion"),
            id,
        ))

        db.commit()
        cur.close()
        return {"ok": True}

    except HTTPException:
        db.rollback()
        raise

    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


# =====================================================
# TIENDA (PÚBLICA)
# =====================================================

# =====================================================
# TIENDA (PÚBLICA) - FIX DEFINITIVO
# =====================================================

from fastapi import Depends

@app.get("/tienda")
def tienda(categoria: str | None = None, db=Depends(get_db)):
    cur = db.cursor()

    if categoria:
        cur.execute("""
            SELECT
                id,
                nombre,
                stock,
                precio,
                moneda,
                foto_url,
                categoria,
                color,
                bateria,
                condicion,
                precio_revendedor
            FROM productos_tiendaone
            WHERE categoria=%s
              AND foto_url IS NOT NULL
              AND stock > 0
            ORDER BY nombre
        """, (categoria,))
    else:
        cur.execute("""
            SELECT
                id,
                nombre,
                stock,
                precio,
                moneda,
                foto_url,
                categoria,
                color,
                bateria,
                condicion,
                precio_revendedor
            FROM productos_tiendaone
            WHERE foto_url IS NOT NULL
              AND stock > 0
            ORDER BY nombre
        """)

    rows = cur.fetchall()

    productos = [
        {
            "id": r[0],
            "nombre": r[1],
            "stock": r[2],
            "precio": float(r[3]) if r[3] is not None else 0,
            "moneda": r[4],                      # 👈 CLAVE
            "foto_url": r[5],
            "categoria": r[6],
            "color": r[7],
            "bateria": r[8],
            "condicion": r[9],
            "precio_revendedor": float(r[10]) if r[10] is not None else 0,
        }
        for r in rows
    ]

    cur.execute("""
        SELECT DISTINCT categoria
        FROM productos_tiendaone
        WHERE categoria IS NOT NULL
        ORDER BY categoria
    """)
    categorias = [r[0] for r in cur.fetchall()]

    cur.close()

    return {
        "productos": productos,
        "categorias": categorias,
        "categoria_seleccionada": categoria,
    }


# =====================================================
# EXPORTAR STOCK
# =====================================================

from openpyxl.styles import Font, Alignment, PatternFill
from fastapi.responses import StreamingResponse

@app.get("/exportar_stock")
def exportar_stock(db=Depends(get_db)):
    cur = db.cursor()
    cur.execute("""
        SELECT id, nombre, codigo_barras, num, color, bateria,
               condicion, stock, precio, precio_costo, precio_revendedor
        FROM productos_tiendaone
        ORDER BY nombre
    """)
    productos = cur.fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Stock"

    headers = [
        "ID", "Nombre", "Código", "Núm", "Color", "Batería",
        "Condición", "Stock", "Precio Venta", "Precio Costo", "Precio Rev."
    ]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4B5563", end_color="4B5563", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    for p in productos:
        ws.append(list(p))

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = max(
            len(str(c.value)) if c.value else 0 for c in col
        ) + 2

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=stock.xlsx"}
    )


from fastapi import Query, Depends
from datetime import datetime, timedelta

from fastapi import Depends, Query
from datetime import datetime, timedelta

from fastapi import Query, Depends
from datetime import datetime

@app.get("/transacciones")
def listar_transacciones(
    desde: str = Query(...),
    hasta: str = Query(...),
    db=Depends(get_db)
):
    cur = db.cursor()

    desde_dt = datetime.fromisoformat(desde)
    hasta_dt = datetime.fromisoformat(hasta)

    # =====================================================
    # 🛒 VENTAS
    # =====================================================
    cur.execute("""
        SELECT
            id,
            fecha,
            dni_cliente,
            total
        FROM ventas_tiendaone
        WHERE fecha BETWEEN %s AND %s
        ORDER BY fecha DESC
    """, (desde_dt, hasta_dt))

    ventas_raw = cur.fetchall()
    ventas = []

    for v in ventas_raw:
        venta_id = v[0]

        # -------- ITEMS --------
        cur.execute("""
            SELECT
                COALESCE(p.nombre, vi.nombre_manual) AS producto,
                vi.cantidad,
                vi.precio_unitario,
                vi.moneda,
                vi.total,
                vi.tipo_precio
            FROM ventas_items_tiendaone vi
            LEFT JOIN productos_tiendaone p ON p.id = vi.producto_id
            WHERE vi.venta_id = %s
        """, (venta_id,))

        items = [
            {
                "producto": r[0],
                "cantidad": r[1],
                "precio_unitario": float(r[2]),
                "moneda": r[3],
                "total": float(r[4]),
                "tipo_precio": r[5],
            }
            for r in cur.fetchall()
        ]

        # -------- PAGOS --------
        cur.execute("""
            SELECT metodo, moneda, monto
            FROM pagos_tiendaone
            WHERE venta_id = %s
        """, (venta_id,))

        pagos = [
            {
                "metodo": p[0],
                "moneda": p[1],
                "monto": float(p[2]),
            }
            for p in cur.fetchall()
        ]

        ventas.append({
            "tipo": "venta",
            "id": venta_id,
            "fecha": v[1],
            "dni_cliente": v[2],
            "total": float(v[3]),
            "items": items,
            "pagos": pagos,
        })

    # =====================================================
    # 🔧 REPARACIONES
    # =====================================================
    cur.execute("""
        SELECT
            r.id,
            r.fecha,
            r.cliente,
            r.equipo,
            r.descripcion,
            p.metodo,
            p.moneda,
            p.monto
        FROM pagos_tiendaone p
        JOIN reparaciones_tiendaone r ON r.id = p.reparacion_id
        WHERE r.fecha BETWEEN %s AND %s
        ORDER BY r.fecha DESC
    """, (desde_dt, hasta_dt))

    reparaciones = [
        {
            "tipo": "reparacion",
            "id": r[0],
            "fecha": r[1],
            "cliente": r[2],
            "equipo": r[3],
            "reparacion": r[4],
            "total": float(r[7]),
            "pagos": [
                {
                    "metodo": r[5],
                    "moneda": r[6],
                    "monto": float(r[7]),
                }
            ],
        }
        for r in cur.fetchall()
    ]

    cur.close()

    # =====================================================
    # 📦 RESPUESTA UNIFICADA
    # =====================================================
    return {
        "ventas": ventas,
        "reparaciones": reparaciones
    }





import pandas as pd
from fastapi.responses import StreamingResponse
from io import BytesIO

from fastapi import Depends
from fastapi.responses import StreamingResponse
from datetime import datetime
from io import BytesIO
import pandas as pd

@app.get("/transacciones/exportar")
def exportar_transacciones(
    desde: str,
    hasta: str,
    db=Depends(get_db)
):
    cur = db.cursor()

    desde_dt = datetime.fromisoformat(desde)
    hasta_dt = datetime.fromisoformat(hasta)

    # =====================================================
    # 🛒 VENTAS (MISMO FILTRO QUE LA PANTALLA)
    # =====================================================
    cur.execute("""
        SELECT
            v.id AS venta_id,
            v.fecha,
            v.dni_cliente,

            COALESCE(p.nombre, vi.nombre_manual) AS producto,
            vi.cantidad,
            vi.precio_unitario,
            vi.moneda AS moneda_item,
            vi.total AS total_item,
            vi.tipo_precio,

            string_agg(
                pg.metodo || ' ' || pg.moneda || ' ' || pg.monto,
                ' | '
            ) AS pagos
        FROM ventas_tiendaone v
        JOIN ventas_items_tiendaone vi
            ON vi.venta_id = v.id
        LEFT JOIN productos_tiendaone p
            ON p.id = vi.producto_id
        LEFT JOIN pagos_tiendaone pg
            ON pg.venta_id = v.id
        WHERE v.fecha BETWEEN %s AND %s
          AND v.anulada = false
        GROUP BY
            v.id, v.fecha, v.dni_cliente,
            p.nombre, vi.nombre_manual,
            vi.cantidad, vi.precio_unitario,
            vi.moneda, vi.total, vi.tipo_precio
        ORDER BY v.fecha DESC, v.id
    """, (desde_dt, hasta_dt))

    ventas_rows = cur.fetchall()
    ventas_cols = [c.name for c in cur.description]
    df_ventas = pd.DataFrame(ventas_rows, columns=ventas_cols)

    # =====================================================
    # 🔧 REPARACIONES
    # =====================================================
    cur.execute("""
        SELECT
            r.id AS reparacion_id,
            r.fecha,
            r.cliente,
            r.equipo,
            r.reparacion,
            r.total,

            string_agg(
                pr.metodo || ' ' || pr.moneda || ' ' || pr.monto,
                ' | '
            ) AS pagos
        FROM reparaciones r
        LEFT JOIN pagos_reparaciones pr
            ON pr.reparacion_id = r.id
        WHERE r.fecha BETWEEN %s AND %s
        GROUP BY r.id, r.fecha, r.cliente, r.equipo, r.reparacion, r.total
        ORDER BY r.fecha DESC, r.id
    """, (desde_dt, hasta_dt))

    rep_rows = cur.fetchall()
    rep_cols = [c.name for c in cur.description]
    df_reparaciones = pd.DataFrame(rep_rows, columns=rep_cols)

    cur.close()

    # =====================================================
    # 📤 EXPORTAR EXCEL (2 HOJAS)
    # =====================================================
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_ventas.to_excel(
            writer,
            sheet_name="Ventas",
            index=False
        )
        df_reparaciones.to_excel(
            writer,
            sheet_name="Reparaciones",
            index=False
        )

    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition":
                "attachment; filename=transacciones_detalladas.xlsx"
        }
    )



from fastapi import HTTPException

from fastapi import HTTPException

@app.post("/transacciones/{venta_id}/anular")
def eliminar_transaccion(
    venta_id: int,
    db = Depends(get_db)
):
    cur = db.cursor()

    # Buscar venta
    cur.execute("""
        SELECT
            id,
            producto_id,
            cantidad
        FROM ventas_tiendaone
        WHERE id = %s
    """, (venta_id,))

    venta = cur.fetchone()

    if not venta:
        cur.close()
        raise HTTPException(status_code=404, detail="Transacción no encontrada")

    try:
        # 👉 Si es producto, devolver stock
        if venta["producto_id"] is not None:
            cur.execute("""
                UPDATE productos_tiendaone
                SET stock = stock + %s
                WHERE id = %s
            """, (venta["cantidad"], venta["producto_id"]))

        # 👉 Eliminar transacción
        cur.execute("""
            DELETE FROM ventas_tiendaone
            WHERE id = %s
        """, (venta_id,))

        db.commit()
        return {"ok": True}

    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

    finally:
        cur.close()

from fastapi import HTTPException, Depends
from psycopg2.extras import DictCursor
from datetime import datetime

ESTADOS_VALIDOS = (
    "ingresado",
    "en_reparacion",
    "listo",
    "retirado",
)

@app.post("/reparaciones")
def crear_reparacion(data: dict, db=Depends(get_db)):
    cur = db.cursor(cursor_factory=DictCursor)

    try:
        estado = data.get("estado", "ingresado")
        if estado not in ESTADOS_VALIDOS:
            raise HTTPException(400, f"Estado inválido: {estado}")

        cur.execute("""
            INSERT INTO reparaciones_tiendaone (
                descripcion,
                cantidad,
                precio,
                estado,
                cobrada,
                tipo_pago,
                cliente,
                telefono,
                equipo,
                imei,
                fecha,
                created_at
            ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,NOW(),NOW())
            RETURNING id
        """, (
            data["reparacion"],
            1,
            float(data["precio"]),
            estado,
            False,
            "pendiente",          # 👈 CLAVE
            data.get("cliente"),
            data.get("telefono"),
            data.get("equipo"),
            data.get("imei"),
        ))

        reparacion_id = cur.fetchone()["id"]
        db.commit()
        return {"ok": True, "id": reparacion_id}

    except Exception as e:
        db.rollback()
        print("ERROR CREAR REPARACION:", e)
        raise HTTPException(500, str(e))

    finally:
        cur.close()





@app.get("/reparaciones")
def listar_reparaciones(estado: str | None = None, db=Depends(get_db)):
    cur = db.cursor(cursor_factory=DictCursor)

    if estado:
        cur.execute("""
            SELECT
                id,
                cliente,
                dni,
                telefono,
                equipo,
                imei,
                descripcion AS reparacion,
                precio,
                estado,
                cobrada
            FROM reparaciones_tiendaone
            WHERE estado = %s
            ORDER BY created_at DESC
        """, (estado,))
    else:
        cur.execute("""
            SELECT
                id,
                cliente,
                dni,
                telefono,
                equipo,
                imei,
                descripcion AS reparacion,
                precio,
                estado,
                cobrada
            FROM reparaciones_tiendaone
            ORDER BY created_at DESC
        """)

    rows = cur.fetchall()
    cur.close()

    # 🔥 CLAVE: convertir a dict real
    return [dict(r) for r in rows]



@app.patch("/reparaciones/{id}/estado")
def cambiar_estado_reparacion(id: int, data: dict, db=Depends(get_db)):
    estado = data.get("estado")
    if estado not in ESTADOS_VALIDOS:
        raise HTTPException(400, "Estado inválido")

    cur = db.cursor()
    cur.execute("""
        UPDATE reparaciones_tiendaone
        SET estado = %s
        WHERE id = %s
    """, (estado, id))

    if cur.rowcount == 0:
        db.rollback()
        raise HTTPException(404, "Reparación no encontrada")

    db.commit()
    cur.close()
    return {"ok": True}

@app.post("/reparaciones/{rep_id}/cobrar")
def cobrar_reparacion(rep_id: int, data: dict, db=Depends(get_db)):
    cur = db.cursor()

    cur.execute("""
        UPDATE reparaciones_tiendaone
        SET
            cobrada = TRUE,
            estado = 'retirado',
            fecha = NOW()
        WHERE id = %s
    """, (rep_id,))

    cur.execute("""
        INSERT INTO pagos_tiendaone (
            venta_id,
            metodo,
            moneda,
            monto,
            reparacion_id
        )
        VALUES (
            NULL,
            %s,
            %s,
            %s,
            %s
        )
    """, (
        data["metodo"],
        data["moneda"],
        data["monto"],
        rep_id
    ))

    db.commit()
    cur.close()

    return {"ok": True}





@app.delete("/reparaciones/{id}")
def eliminar_reparacion(id: int, db=Depends(get_db)):
    cur = db.cursor()
    cur.execute("DELETE FROM reparaciones_tiendaone WHERE id = %s", (id,))
    if cur.rowcount == 0:
        db.rollback()
        raise HTTPException(404, "Reparación no encontrada")
    db.commit()
    cur.close()
    return {"ok": True}




if __name__ == '__main__':
    import os
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port)
